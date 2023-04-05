from tkinter import *
import openpyxl

# Crear la ventana
ventana = Tk()
ventana.geometry("300x250")
ventana.title("Paysup")
ventana.iconbitmap("paysup.ico") # Agregar el icono

# Cambiar los colores de la interfaz
ventana.configure(bg="#F8E9A1")
color_titulo = "#0A3D62"
color_botones = "#328CC1"

# Función para guardar la información en el archivo de Excel
def guardar_excel():
    # Abrir el archivo de Excel o crear uno nuevo si no existe
    try:
        libro = openpyxl.load_workbook('propinas.xlsx')
    except FileNotFoundError:
        libro = openpyxl.Workbook()
    
    # Seleccionar la hoja de cálculo
    hoja = libro.active
    
    # Obtener los datos de la calculadora
    subtotal = float(subtotal_var.get())
    propina = float(propina_var.get())
    total = subtotal + (subtotal * propina / 100)
    
    # Añadir los datos a la hoja de cálculo
    fila = hoja.max_row + 1
    hoja.cell(row=fila, column=1, value=subtotal)
    hoja.cell(row=fila, column=2, value=propina)
    hoja.cell(row=fila, column=3, value=total)
    
    # Guardar el archivo
    libro.save('propinas.xlsx')
    
    # Limpiar los campos de la calculadora
    subtotal_var.set("")
    propina_var.set("")
    total_label.config(text="")
    
# Función para calcular la propina y el total
def calcular():
    # Obtener los datos de la calculadora
    subtotal = float(subtotal_var.get())
    propina = float(propina_var.get())
    
    # Calcular la propina y el total
    total_propina = subtotal * propina / 100
    total = subtotal + total_propina
    
    # Mostrar los resultados en la interfaz
    total_label.config(text="Total: ${:.2f} (Propina: ${:.2f})".format(total, total_propina))

# Variables de control para los campos de entrada
subtotal_var = StringVar()
propina_var = StringVar()

# Etiquetas
subtotal_label = Label(ventana, text="Subtotal:", bg="#F8E9A1", fg=color_titulo)
subtotal_label.pack()
propina_label = Label(ventana, text="Propina (%):", bg="#F8E9A1", fg=color_titulo)
propina_label.pack()

# Campos de entrada
subtotal_entry = Entry(ventana, textvariable=subtotal_var)
subtotal_entry.pack()
propina_entry = Entry(ventana, textvariable=propina_var)
propina_entry.pack()

# Botones
calcular_button = Button(ventana, text="Calcular", command=calcular, bg=color_botones)
calcular_button.pack()
guardar_button = Button(ventana, text="Guardar en Excel", command=guardar_excel, bg=color_botones)
guardar_button.pack()

# Etiqueta para mostrar el total y la propina
total_label = Label(ventana, text="", bg="#F8E9A1", fg=color_titulo)
total_label.pack()

# Iniciar la ventana
ventana.mainloop()
